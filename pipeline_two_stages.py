#%%
from datetime import datetime
import uuid
from pathlib import Path 
from typing import Optional

import pandas as pd
import numpy as np
import hashlib
import os

from src.extractor import (read_statements,
                            get_file_hash, 
                            log_processed_file,
                            get_processed_files,
                            ingest_data)  
from src.logic import (clean_and_transform_scotia_cc, 
                        clean_and_transform_ws_cheque_acc, 
                        clean_and_transform_ws_cc, 
                        normalize_merchant_names
                        )
from models.categorize import (categorize_transactions,
                        mapping_transactions_dict,
                        get_llm_suggestions,
                        scrub_transaction,
                        review_and_update, 
                        cat_subs_map
                        )
# -------------------------------------------------------------------
# Name to Transformer Mapping
# -------------------------------------------------------------------
name2transformer = {
    "WS_cc": clean_and_transform_ws_cc,
    "SC_cc": clean_and_transform_scotia_cc,
    "WS_cheque": clean_and_transform_ws_cheque_acc,
}

import glob

raw_folder = "data/raw"
log_file = "data/processed_log.txt"
raw_path = 'data/all_raw_transactions.parquet'
processed_hashes = get_processed_files()

# STAGE 1: INGESTION (The "Incremental Update")
csv_files = glob.glob(f"{raw_folder}/**/*.csv", recursive=True)

for file_path in csv_files:
    for key, transformer in name2transformer.items():
        if key in file_path:
            ingest_data(file_path, processed_hashes, transformer, raw_path)
            break

# STAGE 2: TRANSFORMATION (The "Full Refresh")
# This applies your LATEST logic to your ENTIRE history
if os.path.exists(raw_path):
    try:
        print("Refreshing master file with latest transformation logic...")
        
        # 1. Load the entire historical raw data
        full_history = pd.read_parquet(raw_path)
        
        # 2. Apply your latest categorization and cleaning rules
        # This ensures your 'bug fixes' apply to old rows too
        master_df = categorize_transactions (full_history, mapping_transactions_dict) 
        
        # 3. Overwrite the final processed file
        master_df.to_parquet('data/processed/master_transactions.parquet', index=False)
        print(f"Success! Master file contains {len(master_df)} categorized transactions.") 
    except Exception as e:
        print(f"Failed to refresh master file: {e}")
else: 
    print(f"No raw data found at {raw_path} to transform.")

    
# STAGE 3: CATEGORIZATION (THE "FULL REFRESH")
# Initial Categorization
master_df = pd.read_parquet('data/processed/master_transactions.parquet')
master_df = categorize_transactions (master_df, mapping_transactions_dict)

# Filter for LLM processing
uncategorized_mask = (master_df['category'] == 'Uncategorized') & (master_df['subcategory'] == 'Other')
uncategorized_items = master_df[uncategorized_mask]

if not uncategorized_items.empty:
    print(f"Requesting LLM suggestions for {len(uncategorized_items)} uncategorized items...")
    llm_results = get_llm_suggestions(uncategorized_items)
    
    # Review and Update Master Dataframe
    master_df = review_and_update(master_df, llm_results)

# Final Result
print("\nWorkflow complete. Master Dataframe ready for reporting.")

# Reporting
import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment

budgets_df = pd.read_csv('models/budgets.csv')

def export_financial_report(master_df, budgets_df, filename="financial_summary_report.xlsx"):
    # Dynamic Month Detection
    latest_period = master_df.sort_values(['year', 'month']).iloc[-1]
    target_month, target_year = int(latest_period['month']), int(latest_period['year'])

    # --- 1. SHEET 1: MONTHLY CATEGORY SPENDING ---
    m_actual = master_df[(master_df['year'] == target_year) & (master_df['month'] == target_month)]
    m_budget = budgets_df[(budgets_df['year'] == target_year) & (budgets_df['month'] == target_month)]
    
    # Aggregation for current month
    cat_spend = m_actual[(m_actual['amount'] < 0) & \
                         (m_actual['category'] != 'Income') & \
                         (m_actual['category'] != 'Transfer')]\
                            .groupby('category')['amount'].sum().abs().reset_index()
    cat_budget = m_budget[(m_budget['category'] != 'Income') & \
                          (m_budget['category'] != 'Transfer')]\
                            .groupby('category')['amount'].sum().reset_index()
    
    # Aggregation for YTD Trends
    ytd_act = master_df[(master_df['year'] == target_year) & \
                     (master_df['month'] <= target_month) & \
                     (master_df['category'] != 'Income') & \
                     (master_df['category'] != 'Transfer')]
    ytd_bud = budgets_df[(budgets_df['year'] == target_year) & \
                         (budgets_df['month'] <= target_month) & \
                         (budgets_df['category'] != 'Income') & \
                         (budgets_df['category'] != 'Transfer')]
    ytd_spend_map = ytd_act[ytd_act['amount'] < 0].groupby('category')['amount'].sum().abs()
    ytd_budget_map = ytd_bud.groupby('category')['amount'].sum()

    report_1 = pd.merge(cat_budget, cat_spend, on='category', how='left', suffixes=('_b', '_s')).fillna(0)
    report_1['Monthly Trend'] = np.where(report_1['amount_s'] > report_1['amount_b'], 'Overspent', 'Underspent')
    report_1['Accumulative Trend YTD'] = report_1['category'].apply(lambda x: 'Overspent' if ytd_spend_map.get(x,0) > ytd_budget_map.get(x,0) else 'Underspent')
    
    # New Column: % spending ratio (numeric for Excel casting)
    report_1['% spending ratio'] = np.where(report_1['amount_b'] != 0, report_1['amount_s'] / report_1['amount_b'], 0)
    
    # Sort: Overspent first
    report_1['Monthly Trend'] = pd.Categorical(report_1['Monthly Trend'], categories=['Overspent', 'Underspent'], ordered=True)
    report_1 = report_1.sort_values(['Monthly Trend', 'amount_s'], ascending=[True, False])
    
    sheet1_df = report_1[['category', 'amount_b', 'amount_s', '% spending ratio', 'Monthly Trend', 'Accumulative Trend YTD']]
    sheet1_df.columns = ['Category', 'Budget', 'Spending', '% spending ratio', 'Monthly Trend', 'Accumulative Trend YTD']

    # --- 2. SHEET 2: TOP 5 NON-RECURRING OF LAST MONTH ---
    # Logic: Dynamic last month + is_recurring == False
    m_non_rec = m_actual[(m_actual['is_recurring'] == False) & \
                        (m_actual['amount'] < 0) & \
                        (m_actual['transaction_type'] == 'Expense')]
    sheet2_df = m_non_rec.groupby('description')['amount'].sum().abs().nlargest(5).reset_index()
    sheet2_df.columns = ['Description', 'Total Value']

    # --- 3. SHEET 3: YEARLY FINANCIAL SUMMARY ---
    summary = []
    for yr, group in master_df.groupby('year'):
        inc = group[group['amount'] > 0]['amount'].sum()
        exp = abs(group[group['amount'] < 0]['amount'].sum())
        label = str(yr) if group['month'].nunique() == 12 else f"{yr} (YTD)"
        summary.append({'Year Period': label, 'Net Income': inc, 'Net Expenditures': exp, 'Savings': inc - exp})
    sheet3_df = pd.DataFrame(summary)

    # --- 4. EXCEL EXPORT WITH NUMBER FORMATS ---
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        sheet1_df.to_excel(writer, sheet_name='Monthly Spending', index=False)
        sheet2_df.to_excel(writer, sheet_name='Top 5 Non-Recurring', index=False)
        sheet3_df.to_excel(writer, sheet_name='Yearly Summary', index=False)

        # Standard Accounting and Percentage formats for Excel
        accounting_fmt = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
        percent_fmt = '0%'
        
        for sheetname in writer.sheets:
            ws = writer.sheets[sheetname]
            
            # Format Headers
            for cell in ws[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Apply Data Types
            if sheetname == 'Monthly Spending':
                for row in ws.iter_rows(min_row=2):
                    row[1].number_format = accounting_fmt # Budget column
                    row[2].number_format = accounting_fmt # Spending column
                    row[3].number_format = percent_fmt    # % spending ratio
            
            elif sheetname == 'Top 5 Non-Recurring':
                for row in ws.iter_rows(min_row=2):
                    row[1].number_format = accounting_fmt # Total Value column
            
            elif sheetname == 'Yearly Summary':
                for row in ws.iter_rows(min_row=2):
                    row[1].number_format = accounting_fmt # Net Income
                    row[2].number_format = accounting_fmt # Net Expenditures
                    row[3].number_format = accounting_fmt # Savings

            # Set Column Widths
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 25

    print(f"Excel report saved as: {filename}")

# Usage: 
export_financial_report(master_df, budgets_df)
# %%
